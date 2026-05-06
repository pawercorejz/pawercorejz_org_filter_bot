import os
import re
import asyncio
import tempfile
from openpyxl import load_workbook, Workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

BOT_TOKEN = os.getenv("BOT_TOKEN")

user_queues = {}
user_tasks = {}


def progress_bar(percent):
    filled = int(percent / 10)
    empty = 10 - filled
    return "█" * filled + "░" * empty


def normalize_phone(value):
    if value is None:
        return ""

    digits = re.sub(r"\D", "", str(value))

    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]

    if len(digits) == 10:
        digits = "7" + digits

    return digits


def load_blacklist():
    blacklist = set()

    if os.path.exists("blacklist.txt"):
        with open("blacklist.txt", "r", encoding="utf-8") as f:
            for line in f:
                phone = normalize_phone(line.strip())
                if phone:
                    blacklist.add(phone)

    return blacklist


def is_person(fullname):
    if not fullname:
        return False

    text = str(fullname).strip()
    lower = text.lower()
    parts = text.split()

    if len(parts) != 3:
        return False

    bad_words = [
        "ооо", "оао", "ао", "зао", "пао", "ип",
        "компания", "управляющая", "организация",
        "администрация", "отдел", "отделение", "управление",
        "участок", "полиция", "суд", "банк", "школа",
        "лицей", "гимназия", "университет", "институт",
        "колледж", "центр", "служба", "фонд", "завод",
        "фабрика", "предприятие", "общество", "товарищество",
        "кооператив", "агентство", "министерство",
        "департамент", "комитет", "больница", "поликлиника",
        "клиника", "аптека", "магазин", "салон", "студия",
        "кафе", "ресторан", "группа", "холдинг", "жкх",
        "тсж", "жск", "снт",

        "федерация", "училище", "мтц", "энергетика",
        "комплекс", "гимнастика", "онкодиспансер",
        "рдз", "комиссия", "фбуагентство", "транспорта",
        "металлистизделие", "лаборатория", "хроматографии",
        "терминал", "многофункциональный", "мазутный",
        "музыкальное", "художественной", "апелляционная",
        "тверской", "северное", "тушино", "промышленная",

        "джет", "сервис", "мир", "рти", "петербург",
        "алюминстрой", "рекламные", "материалы",
        "монтажника", "ммуc", "ммус", "контур", "фокус",
        "кузовной", "цех", "алмас", "лотос", "действует",
        "медиа", "бизнес", "консалтинг", "цирк", "вывоз",
        "мусора", "долгопрудный", "маллиотт", "бульвар",
        "отель", "команда", "нпо", "монолит", "миэль",
        "курской", "транс", "лига", "мгу", "ломоносова",

        "минздрав", "правительство", "министерство",
        "хозяйства", "комбинат", "техникум", "филиал",
        "дирекция", "строительства", "ветеринарная", "станция"
    ]

    for word in bad_words:
        if word in lower:
            return False

    for part in parts:
        if not re.match(r"^[А-ЯЁ][а-яё-]+$", part):
            return False

    return True


def process_excel(input_path, output_path):
    blacklist = load_blacklist()

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]

    try:
        request_col = headers_lower.index("request")
        fullname_col = headers_lower.index("fullname")
    except ValueError:
        raise Exception("Не нашёл столбцы request и FullName")

    address_col = headers_lower.index("address") if "address" in headers_lower else None

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "result"
    new_ws.append(["request", "FullName", "Address"])

    total = 0
    kept = 0
    removed = 0

    for row in rows:
        total += 1

        number = normalize_phone(row[request_col])
        fullname = row[fullname_col]

        if number in blacklist:
            removed += 1
            continue

        if not (is_person(fullname) or fullname is None or str(fullname).strip() == ""):
            removed += 1
            continue

        address_value = row[address_col] if address_col is not None else ""

        new_ws.append([
            row[request_col],
            row[fullname_col],
            address_value
        ])

        kept += 1

    new_wb.save(output_path)

    percent = round((removed / total) * 100, 2) if total else 0

    return kept, removed, total, percent


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🚀 Я готов\n\n"
        "Кидай Excel файлы (.xlsx)\n\n"
        "🔹 Удаляю организации\n"
        "🔹 Удаляю номера из blacklist.txt\n"
        "🔹 Оставляю ФИО и пустые FullName\n"
        "🔹 Показываю красивую статистику"
    )


async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document

    if not document.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("⚠️ Только .xlsx файл")
        return

    user_id = update.effective_user.id

    if user_id not in user_queues:
        user_queues[user_id] = []

    user_queues[user_id].append({
        "document": document,
        "chat_id": update.effective_chat.id
    })

    await update.message.reply_text(f"📥 Файл принят ✅\n{document.file_name}")

    if user_id not in user_tasks or user_tasks[user_id].done():
        user_tasks[user_id] = asyncio.create_task(process_queue(context, user_id))


async def process_queue(context: ContextTypes.DEFAULT_TYPE, user_id: int):
    await asyncio.sleep(3)

    queue = user_queues.get(user_id, [])
    total_files = len(queue)

    if total_files == 0:
        return

    chat_id = queue[0]["chat_id"]
    file_index = 0

    while user_queues.get(user_id):
        item = user_queues[user_id].pop(0)
        document = item["document"]
        file_index += 1

        progress = await context.bot.send_message(
            chat_id=chat_id,
            text=(
                f"🚀 Запускаю обработку\n"
                f"{progress_bar(0)} 0%\n\n"
                f"📄 Файл {file_index} из {total_files}\n"
                f"{document.file_name}"
            )
        )

        input_path = None
        output_path = None

        try:
            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"📥 Скачиваю файл...\n"
                f"{progress_bar(15)} 15%\n\n"
                f"📄 {document.file_name}"
            )

            tg_file = await document.get_file()

            safe_name = document.file_name.replace("/", "_").replace("\\", "_")
            input_path = os.path.join(tempfile.gettempdir(), safe_name)
            output_name = safe_name.replace(".xlsx", "_filtered.xlsx")
            output_path = os.path.join(tempfile.gettempdir(), output_name)

            await tg_file.download_to_drive(input_path)

            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"🧠 Анализирую структуру файла...\n"
                f"{progress_bar(30)} 30%\n\n"
                f"📄 {document.file_name}"
            )

            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"🔍 Проверяю blacklist...\n"
                f"{progress_bar(45)} 45%\n\n"
                f"📄 {document.file_name}"
            )

            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"🧹 Удаляю организации и лишние строки...\n"
                f"{progress_bar(65)} 65%\n\n"
                f"📄 {document.file_name}"
            )

            kept, removed, total, percent = process_excel(input_path, output_path)

            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"📊 Считаю статистику...\n"
                f"{progress_bar(85)} 85%\n\n"
                f"📄 {document.file_name}"
            )

            caption = (
                f"✅ Файл обработан\n\n"
                f"📄 Файл: {document.file_name}\n"
                f"📊 Всего строк: {total}\n\n"
                f"🟢 Осталось: {kept}\n"
                f"🔴 Удалено: {removed}\n"
                f"📉 Очистка: {percent}%"
            )

            await asyncio.sleep(0.4)
            await progress.edit_text(
                f"📤 Отправляю результат...\n"
                f"{progress_bar(95)} 95%\n\n"
                f"📄 {document.file_name}"
            )

            with open(output_path, "rb") as f:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=f,
                    filename=output_name,
                    caption=caption
                )

            await progress.edit_text(
                f"✅ Готово\n"
                f"{progress_bar(100)} 100%\n\n"
                f"📄 {document.file_name}"
            )

        except Exception as e:
            await progress.edit_text(
                f"❌ Ошибка обработки\n\n"
                f"📄 {document.file_name}\n"
                f"⚠️ {e}"
            )

        finally:
            if input_path and os.path.exists(input_path):
                os.remove(input_path)
            if output_path and os.path.exists(output_path):
                os.remove(output_path)

    await context.bot.send_message(
        chat_id=chat_id,
        text="✅ Все файлы обработаны"
    )


def main():
    if not BOT_TOKEN:
        raise ValueError("BOT_TOKEN не найден")

    print("Бот запускается...")

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))

    print("Бот запущен!")

    app.run_polling()


if __name__ == "__main__":
    main()
